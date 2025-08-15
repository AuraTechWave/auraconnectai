# backend/modules/analytics/services/export_service.py

import io
import csv
import logging
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
from decimal import Decimal
import tempfile
import os

from sqlalchemy.orm import Session
from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse

from ..schemas.analytics_schemas import (
    SalesReportRequest,
    SalesDetailResponse,
    StaffPerformanceResponse,
    ProductPerformanceResponse,
    SalesSummaryResponse,
)
from .sales_report_service import SalesReportService

logger = logging.getLogger(__name__)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("ReportLab not available. PDF export functionality will be limited.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning(
        "OpenPyXL not available. Excel export functionality will be limited."
    )


class ExportService:
    """Service for exporting sales reports in various formats"""

    def __init__(self, db: Session):
        self.db = db
        self.sales_service = SalesReportService(db)

    async def export_sales_report(
        self, request: SalesReportRequest, format_type: str, executed_by: int
    ) -> StreamingResponse:
        """
        Export sales report in the specified format.

        Args:
            request: Sales report request with filters and configuration
            format_type: Export format (csv, pdf, xlsx)
            executed_by: ID of the user requesting the export

        Returns:
            StreamingResponse with the exported file
        """
        try:
            # Generate the report data based on type
            if request.report_type.value == "sales_detailed":
                data = self.sales_service.generate_detailed_sales_report(
                    request.filters,
                    page=1,
                    per_page=10000,  # Large limit for export
                    sort_by=request.sort_by,
                    sort_order=request.sort_order,
                )
                export_data = data.items

            elif request.report_type.value == "staff_performance":
                export_data = self.sales_service.generate_staff_performance_report(
                    request.filters, page=1, per_page=1000
                )

            elif request.report_type.value == "product_performance":
                export_data = self.sales_service.generate_product_performance_report(
                    request.filters, page=1, per_page=1000
                )

            else:  # sales_summary
                summary_data = self.sales_service.generate_sales_summary(
                    request.filters
                )
                export_data = [summary_data]

            # Export based on format
            if format_type.lower() == "csv":
                return await self._export_csv(export_data, request.report_type.value)
            elif format_type.lower() == "pdf":
                return await self._export_pdf(
                    export_data, request.report_type.value, request
                )
            elif format_type.lower() == "xlsx":
                return await self._export_excel(
                    export_data, request.report_type.value, request
                )
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Unsupported export format: {format_type}",
                )

        except Exception as e:
            logger.error(f"Error exporting sales report: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to export sales report",
            )

    async def _export_csv(
        self,
        data: List[
            Union[
                SalesDetailResponse,
                StaffPerformanceResponse,
                ProductPerformanceResponse,
                SalesSummaryResponse,
            ]
        ],
        report_type: str,
    ) -> StreamingResponse:
        """Export data to CSV format"""

        output = io.StringIO()

        if not data:
            # Return empty CSV with headers
            writer = csv.writer(output)
            writer.writerow(self._get_csv_headers(report_type))
            output.seek(0)

            return StreamingResponse(
                io.BytesIO(output.getvalue().encode("utf-8")),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                },
            )

        # Determine headers based on data type
        headers = self._get_csv_headers(report_type)
        writer = csv.writer(output)
        writer.writerow(headers)

        # Write data rows
        for item in data:
            row = self._format_csv_row(item, report_type)
            writer.writerow(row)

        output.seek(0)

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            },
        )

    async def _export_pdf(
        self, data: List[Any], report_type: str, request: SalesReportRequest
    ) -> StreamingResponse:
        """Export data to PDF format"""

        if not REPORTLAB_AVAILABLE:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="PDF export requires ReportLab library. Please install: pip install reportlab",
            )

        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        temp_path = temp_file.name
        temp_file.close()

        try:
            # Create PDF document
            doc = SimpleDocTemplate(temp_path, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            # Title
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=16,
                spaceAfter=30,
                alignment=1,  # Center alignment
            )
            title = Paragraph(
                f"Sales Report - {report_type.replace('_', ' ').title()}", title_style
            )
            elements.append(title)

            # Report parameters
            params_text = f"""
            Report Period: {request.filters.date_from} to {request.filters.date_to}<br/>
            Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
            Total Records: {len(data)}
            """
            params = Paragraph(params_text, styles["Normal"])
            elements.append(params)
            elements.append(Spacer(1, 20))

            if not data:
                no_data = Paragraph(
                    "No data available for the selected criteria.", styles["Normal"]
                )
                elements.append(no_data)
            else:
                # Create table
                table_data = []
                headers = self._get_pdf_headers(report_type)
                table_data.append(headers)

                # Add data rows
                for item in data:
                    row = self._format_pdf_row(item, report_type)
                    table_data.append(row)

                # Create and style table
                table = Table(table_data)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                            ("FONTSIZE", (0, 1), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )

                elements.append(table)

            # Build PDF
            doc.build(elements)

            # Read file and return as response
            with open(temp_path, "rb") as f:
                pdf_content = f.read()

            return StreamingResponse(
                io.BytesIO(pdf_content),
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                },
            )

        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    async def _export_excel(
        self, data: List[Any], report_type: str, request: SalesReportRequest
    ) -> StreamingResponse:
        """Export data to Excel format"""

        if not OPENPYXL_AVAILABLE:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="Excel export requires openpyxl library. Please install: pip install openpyxl",
            )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Title row
        title = f"Sales Report - {report_type.replace('_', ' ').title()}"
        ws.merge_cells("A1:F1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center_alignment

        # Parameters row
        params = f"Period: {request.filters.date_from} to {request.filters.date_to} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells("A2:F2")
        ws["A2"] = params
        ws["A2"].alignment = center_alignment

        # Headers (starting from row 4)
        headers = self._get_excel_headers(report_type)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Data rows
        if data:
            for row_idx, item in enumerate(data, 5):  # Start from row 5
                row_data = self._format_excel_row(item, report_type)
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            },
        )

    def _get_csv_headers(self, report_type: str) -> List[str]:
        """Get CSV headers based on report type"""

        if report_type == "sales_detailed":
            return [
                "Date",
                "Staff ID",
                "Staff Name",
                "Product ID",
                "Category",
                "Total Orders",
                "Total Revenue",
                "Items Sold",
                "Average Order Value",
                "Discounts",
                "Tax",
                "Net Revenue",
                "Customers",
            ]
        elif report_type == "staff_performance":
            return [
                "Staff ID",
                "Staff Name",
                "Orders Handled",
                "Revenue Generated",
                "Average Order Value",
                "Orders Per Hour",
                "Processing Time",
                "Revenue Rank",
                "Order Rank",
            ]
        elif report_type == "product_performance":
            return [
                "Product ID",
                "Product Name",
                "Category",
                "Quantity Sold",
                "Revenue Generated",
                "Average Price",
                "Order Frequency",
                "Revenue Share %",
                "Quantity Share %",
                "Popularity Rank",
            ]
        else:  # sales_summary
            return [
                "Period Start",
                "Period End",
                "Total Orders",
                "Total Revenue",
                "Items Sold",
                "Average Order Value",
                "Discounts",
                "Tax",
                "Net Revenue",
                "Customers",
                "Revenue Growth %",
                "Order Growth %",
            ]

    def _get_pdf_headers(self, report_type: str) -> List[str]:
        """Get PDF table headers (abbreviated for space)"""

        if report_type == "sales_detailed":
            return ["Date", "Staff", "Orders", "Revenue", "Items", "AOV", "Customers"]
        elif report_type == "staff_performance":
            return ["Staff", "Orders", "Revenue", "AOV", "Rank"]
        elif report_type == "product_performance":
            return ["Product", "Quantity", "Revenue", "Price", "Rank"]
        else:  # sales_summary
            return ["Period", "Orders", "Revenue", "AOV", "Customers", "Growth %"]

    def _get_excel_headers(self, report_type: str) -> List[str]:
        """Get Excel headers (full headers)"""
        return self._get_csv_headers(report_type)

    def _format_csv_row(self, item: Any, report_type: str) -> List[str]:
        """Format data row for CSV export"""

        if report_type == "sales_detailed":
            return [
                str(item.snapshot_date),
                str(item.staff_id or ""),
                item.staff_name or "",
                str(item.product_id or ""),
                item.category_name or "",
                str(item.total_orders),
                str(float(item.total_revenue)),
                str(item.total_items_sold),
                str(float(item.average_order_value)),
                str(float(item.total_discounts)),
                str(float(item.total_tax)),
                str(float(item.net_revenue)),
                str(item.unique_customers),
            ]
        elif report_type == "staff_performance":
            return [
                str(item.staff_id),
                item.staff_name,
                str(item.total_orders_handled),
                str(float(item.total_revenue_generated)),
                str(float(item.average_order_value)),
                str(float(item.orders_per_hour)) if item.orders_per_hour else "",
                (
                    str(float(item.average_processing_time))
                    if item.average_processing_time
                    else ""
                ),
                str(item.revenue_rank or ""),
                str(item.order_count_rank or ""),
            ]
        elif report_type == "product_performance":
            return [
                str(item.product_id),
                item.product_name or "",
                item.category_name or "",
                str(item.quantity_sold),
                str(float(item.revenue_generated)),
                str(float(item.average_price)),
                str(item.order_frequency),
                str(float(item.revenue_share)) if item.revenue_share else "",
                str(float(item.quantity_share)) if item.quantity_share else "",
                str(item.popularity_rank or ""),
            ]
        else:  # sales_summary
            return [
                str(item.period_start),
                str(item.period_end),
                str(item.total_orders),
                str(float(item.total_revenue)),
                str(item.total_items_sold),
                str(float(item.average_order_value)),
                str(float(item.total_discounts)),
                str(float(item.total_tax)),
                str(float(item.net_revenue)),
                str(item.unique_customers),
                str(float(item.revenue_growth)) if item.revenue_growth else "",
                str(float(item.order_growth)) if item.order_growth else "",
            ]

    def _format_pdf_row(self, item: Any, report_type: str) -> List[str]:
        """Format data row for PDF export (abbreviated)"""

        if report_type == "sales_detailed":
            return [
                str(item.snapshot_date),
                item.staff_name or "N/A",
                str(item.total_orders),
                f"${float(item.total_revenue):.2f}",
                str(item.total_items_sold),
                f"${float(item.average_order_value):.2f}",
                str(item.unique_customers),
            ]
        elif report_type == "staff_performance":
            return [
                item.staff_name,
                str(item.total_orders_handled),
                f"${float(item.total_revenue_generated):.2f}",
                f"${float(item.average_order_value):.2f}",
                str(item.revenue_rank or "N/A"),
            ]
        elif report_type == "product_performance":
            return [
                item.product_name or f"Product {item.product_id}",
                str(item.quantity_sold),
                f"${float(item.revenue_generated):.2f}",
                f"${float(item.average_price):.2f}",
                str(item.popularity_rank or "N/A"),
            ]
        else:  # sales_summary
            return [
                f"{item.period_start} to {item.period_end}",
                str(item.total_orders),
                f"${float(item.total_revenue):.2f}",
                f"${float(item.average_order_value):.2f}",
                str(item.unique_customers),
                f"{float(item.revenue_growth):.1f}%" if item.revenue_growth else "N/A",
            ]

    def _format_excel_row(self, item: Any, report_type: str) -> List[Any]:
        """Format data row for Excel export"""

        if report_type == "sales_detailed":
            return [
                item.snapshot_date,
                item.staff_id,
                item.staff_name,
                item.product_id,
                item.category_name,
                item.total_orders,
                float(item.total_revenue),
                item.total_items_sold,
                float(item.average_order_value),
                float(item.total_discounts),
                float(item.total_tax),
                float(item.net_revenue),
                item.unique_customers,
            ]
        elif report_type == "staff_performance":
            return [
                item.staff_id,
                item.staff_name,
                item.total_orders_handled,
                float(item.total_revenue_generated),
                float(item.average_order_value),
                float(item.orders_per_hour) if item.orders_per_hour else None,
                (
                    float(item.average_processing_time)
                    if item.average_processing_time
                    else None
                ),
                item.revenue_rank,
                item.order_count_rank,
            ]
        elif report_type == "product_performance":
            return [
                item.product_id,
                item.product_name,
                item.category_name,
                item.quantity_sold,
                float(item.revenue_generated),
                float(item.average_price),
                item.order_frequency,
                float(item.revenue_share) if item.revenue_share else None,
                float(item.quantity_share) if item.quantity_share else None,
                item.popularity_rank,
            ]
        else:  # sales_summary
            return [
                item.period_start,
                item.period_end,
                item.total_orders,
                float(item.total_revenue),
                item.total_items_sold,
                float(item.average_order_value),
                float(item.total_discounts),
                float(item.total_tax),
                float(item.net_revenue),
                item.unique_customers,
                float(item.revenue_growth) if item.revenue_growth else None,
                float(item.order_growth) if item.order_growth else None,
            ]


# Service factory function
def create_export_service(db: Session) -> ExportService:
    """Create an export service instance"""
    return ExportService(db)
